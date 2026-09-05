import openpyxl

file_path = r'e:\Fee Record Primery 2026.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Aug-26']

# Class definitions and coordinates in Aug-26 sheet
# Fees: PG to 4 = 2000, 5 to 7 = 2500
class_blocks = [
    (69, 12, 'Play Group', 2000),
    (1, 1, 'Nursery', 2000),
    (1, 12, 'Prep', 2000),
    (28, 1, 'One', 2000),
    (28, 12, 'Two', 2000),
    (47, 1, 'Three', 2000),
    (47, 12, 'Four', 2000),
    (69, 1, 'Five', 2500),
    (92, 1, 'Seven', 2500),
]

sql_statements = [
    "-- Migration Script for Tenant 16 (Good Luck School - GLS)",
    "-- Source: August 2026 Fee Record (Aug-26 Sheet)",
    "-- Fees: PG to Class 4 = 2000, Class 5 to 7 = 2500",
    "-- Registration Numbers: Auto-generated sequential GLS-001, GLS-002, ...",
    "START TRANSACTION;\n",
    "SET @tenant_id = 16;\n",
    "-- Clear previous entries for this tenant to avoid duplicate key errors on re-runs",
    "INSERT INTO tenants (id, name, school_name, subdomain, status) VALUES (16, 'Good Luck School', 'Good Luck School', 'goodluck', 'active') ON DUPLICATE KEY UPDATE name=VALUES(name), school_name=VALUES(school_name);",
    "DELETE FROM fee_payments WHERE tenant_id = @tenant_id;",
    "DELETE FROM students WHERE tenant_id = @tenant_id;",
    "DELETE FROM classes WHERE tenant_id = @tenant_id;\n"
]

student_counter = 1
total_payments_inserted = 0
total_paid_sum = 0.0
total_expected_sum = 0.0

for r_start, c_start, class_name, default_fee in class_blocks:
    class_var = f"@class_{class_name.replace(' ', '_').lower()}"
    sql_statements.append(f"\n-- =====================================")
    sql_statements.append(f"-- Class: {class_name} (Monthly Fee: {default_fee})")
    sql_statements.append(f"-- =====================================")
    sql_statements.append(f"INSERT INTO classes (name, default_monthly_fee, tenant_id) VALUES ('{class_name}', {default_fee}, @tenant_id);")
    sql_statements.append(f"SET {class_var} = LAST_INSERT_ID();\n")

    for r in range(r_start + 2, r_start + 25):
        s_no = ws.cell(r, c_start).value
        s_name = ws.cell(r, c_start + 1).value
        prev = ws.cell(r, c_start + 2).value
        fee = ws.cell(r, c_start + 6).value
        paid = ws.cell(r, c_start + 8).value

        s_name_str = str(s_name).strip() if s_name is not None else ''
        s_no_str = str(s_no).strip() if s_no is not None else ''

        # Stop at Total row or empty
        if s_name_str == 'Total' or s_name_str.startswith('CLASS:'):
            break
        if not s_name_str or s_name_str in ['0', '#REF!', 'None', 'nan', 'STUDENT NAME', 'Topic'] or s_no_str.isdigit():
            continue

        reg_no = f"GLS-{student_counter:03d}"
        clean_name = s_name_str.replace("'", "''")
        student_var = f"@student_gls_{student_counter:03d}"

        # Actual tuition fee for student from sheet
        try:
            student_fee = float(fee) if fee is not None else 0.0
        except (ValueError, TypeError):
            student_fee = default_fee

        # In this system, custom_monthly_fee is the CONCESSION DISCOUNT (default_fee - student_fee)
        concession_discount = default_fee - student_fee
        has_concession = 1 if concession_discount > 0 else 0
        total_expected_sum += (default_fee - concession_discount)

        # Previous column handling (outstanding balance / advance)
        notes_parts = []
        prev_balance = 0.0
        try:
            if prev is not None:
                prev_balance = float(prev)
                if prev_balance > 0:
                    notes_parts.append(f"Previous Outstanding: Rs. {int(prev_balance):,}")
                elif prev_balance < 0:
                    notes_parts.append(f"Advance Paid: Rs. {abs(int(prev_balance)):,}")
        except (ValueError, TypeError):
            prev_balance = 0.0

        concession_notes = f"'{', '.join(notes_parts)}'" if notes_parts else "NULL"

        sql_statements.append(f"INSERT INTO students (reg_no, name, class_id, custom_monthly_fee, has_concession, concession_notes, tenant_id) VALUES ('{reg_no}', '{clean_name}', {class_var}, {concession_discount:.2f}, {has_concession}, {concession_notes}, @tenant_id);")
        sql_statements.append(f"SET {student_var} = LAST_INSERT_ID();")

        # August 2026 fee payment insertion (strictly from Paid column in sheet)
        try:
            if paid is not None and float(paid) > 0:
                paid_amount = float(paid)
                sql_statements.append(f"INSERT INTO fee_payments (tenant_id, student_id, month, year, amount_paid, payment_date) VALUES (@tenant_id, {student_var}, 8, 2026, {paid_amount:.2f}, '2026-08-01');")
                total_payments_inserted += 1
                total_paid_sum += paid_amount
        except (ValueError, TypeError):
            pass

        student_counter += 1

sql_statements.append("\nCOMMIT;\n")

target_file = r'd:\school-ms\sql\migrate_tenant_16.sql'
with open(target_file, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_statements))

print(f"Successfully generated clean migration SQL at: {target_file}")
print(f"Total students migrated: {student_counter - 1}")
print(f"Total August 2026 fee payments inserted: {total_payments_inserted}")
print(f"Total Expected Fee Sum: PKR {total_expected_sum:,.2f} (Target: 177,900.00)")
print(f"Total Paid Fee Sum:     PKR {total_paid_sum:,.2f} (Target: 174,700.00)")
